#!/usr/bin/env python3
"""Generate finance import SQL from Cazu_Finanzas.xlsx. Output goes to /tmp (PII)."""
from openpyxl import load_workbook

SRC = "/tmp/cazu-cowork/Projects/Cazú Ceramics/Cazu_Finanzas.xlsx"
OUT = "/tmp/finance-import.sql"
REAL_MONTHS = {"ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO"}


def cents(v):
    if v is None or v == "":
        return "NULL"
    s = str(v).replace("€", "").replace(".", "").replace(",", ".").strip() if isinstance(v, str) else v
    try:
        return str(round(float(s) * 100))
    except (ValueError, TypeError):
        return "NULL"


def q(v):
    if v is None or str(v).strip() == "":
        return "NULL"
    return "'" + str(v).strip().replace("'", "''") + "'"


wb = load_workbook(SRC, data_only=True)
lines = ["-- GENERATED finance import. Apply in Lovable SQL editor. Contains PII.", "BEGIN;"]

# Pagos -> ledger_entries (real months only)
pag = list(wb["Pagos"].iter_rows(values_only=True))[1:]  # skip header
n_pag = 0
for row in pag:
    fecha, mes, alumno, clase, cat, imp, met, estado, cobra, notas = (list(row) + [None] * 10)[:10]
    if mes not in REAL_MONTHS:
        continue
    n_pag += 1
    lines.append(
        "INSERT INTO public.ledger_entries (entry_date, month, student_name, item, category, amount_cents, method, status, notes) VALUES ("
        + f"{q(fecha) if fecha else 'NULL'}, {q(mes)}, {q(alumno)}, {q(clase)}, {q(cat)}, {cents(imp)}, {q(met)}, {q(estado)}, {q(notas)});"
    )

# Gastos -> expense_entries
gas = list(wb["Gastos"].iter_rows(values_only=True))[1:]
n_gas = 0
for row in gas:
    fecha, mes, cat, prov, concepto, imp, met, notas, iva = (list(row) + [None] * 9)[:9]
    if not mes:
        continue
    n_gas += 1
    lines.append(
        "INSERT INTO public.expense_entries (entry_date, month, category, provider, concept, amount_cents, method, notes, vat_cents) VALUES ("
        + f"{q(fecha) if fecha else 'NULL'}, {q(mes)}, {q(cat)}, {q(prov)}, {q(concepto)}, {cents(imp)}, {q(met)}, {q(notas)}, {cents(iva)});"
    )

lines.append("COMMIT;")
lines.append(f"-- pagos rows: {n_pag} | gastos rows: {n_gas}")
open(OUT, "w").write("\n".join(lines))
print(f"Wrote {OUT}: {n_pag} pagos, {n_gas} gastos")
